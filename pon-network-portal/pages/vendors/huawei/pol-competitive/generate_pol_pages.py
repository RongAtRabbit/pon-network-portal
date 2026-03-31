#!/usr/bin/env python3
"""Generate HTML pages from POL competitive analysis Excel file"""

import openpyxl
import os
import html

BASE_DIR = '/home/ryes/.openclaw/workspace/pon-network-portal/pages/vendors/huawei/pol-competitive'
EXCEL_PATH = '/home/ryes/.openclaw/workspace/pon-network-portal/pages/vendors/huawei/pol-competitive/POL竞争分析.xlsx'

# Sheet name to URL slug and display name mapping
SHEET_CONFIG = {
    'PON网络业内竞争': ('01-pon-network-competitor', 'PON网络业内竞争', '🏢'),
    '设备端OLT规格': ('02-olt-specs', '设备端OLT规格', '⚙️'),
    'MA5800规格': ('03-ma5800-specs', 'MA5800规格', '🔌'),
    '方案规格-eSight': ('04-esight-specs', '方案规格-eSight', '💻'),
    '华为产品列表': ('05-huawei-products', '华为产品列表', '📦'),
    '方案规格-eSight（new）': ('06-esight-specs-new', '方案规格-eSight（new）', '💻'),
    '华为POL无线规格(基于WI-FI特性文档）': ('07-pol-wireless-specs', '华为POL无线规格', '📡'),
    '华为ONU规格': ('08-onu-specs', '华为ONU规格', '📡'),
    '华为eSight支持告警': ('09-esight-alarms', '华为eSight支持告警', '⚠️'),
    '标准规格': ('10-standard-specs', '标准规格', '📋'),
    '华为保护组规格': ('11-protection-group', '华为保护组规格', '🛡️'),
    '华为F5G配置流程': ('12-f5g-config', '华为F5G配置流程', '⚡'),
    '无线ONT能力VS 企业AP': ('13-wireless-ont-vs-ap', '无线ONT能力VS企业AP', '📶'),
    'EA5800规格': ('14-ea5800-specs', 'EA5800规格', '🔧'),
    '华为排障指导(tobe fill)': ('15-troubleshooting', '华为排障指导', '🔍'),
}

def escape_html(text):
    if text is None:
        return ''
    return html.escape(str(text))

def row_to_cells(row, max_cols=20):
    """Convert row to HTML table cells"""
    cells = []
    for i, val in enumerate(row):
        if i >= max_cols:
            break
        cells.append(f'<td>{escape_html(val)}</td>')
    return '\n'.join(cells)

def generate_table(ws, max_cols=20):
    """Generate HTML table from worksheet"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return '<p>No data</p>'
    
    # Determine actual column count
    actual_cols = 0
    for row in rows[:5]:
        for i, val in enumerate(row):
            if val is not None and str(val).strip():
                actual_cols = max(actual_cols, i + 1)
    actual_cols = min(actual_cols, max_cols)
    
    html_parts = ['<div class="table-wrapper"><table class="data-table">']
    
    # Header row (first row)
    if rows:
        header_cells = []
        for i, val in enumerate(rows[0][:actual_cols]):
            header_cells.append(f'<th>{escape_html(val)}</th>')
        html_parts.append(f'<thead><tr>{"".join(header_cells)}</tr></thead>')
    
    # Data rows
    html_parts.append('<tbody>')
    for row in rows[1:]:
        if any(c is not None for c in row[:actual_cols]):
            html_parts.append(f'<tr>{row_to_cells(row, actual_cols)}</tr>')
    html_parts.append('</tbody></table></div>')
    
    return '\n'.join(html_parts)

def generate_page_template(title, content, breadcrumb, sheet_info=None):
    """Generate HTML page with consistent template"""
    index_link = '<a href="index.html">行业分析</a>'
    prev_link = sheet_info['prev'] if sheet_info and sheet_info.get('prev') else None
    next_link = sheet_info['next'] if sheet_info and sheet_info.get('next') else None
    
    nav = f'''
        <nav class="navbar">
            <div class="nav-container">
                <a href="../../../index.html" class="nav-logo">
                    <span class="logo-icon">🌐</span>
                    <span class="logo-text">PON Portal</span>
                </a>
                <ul class="nav-menu" id="navMenu">
                    <li><a href="../../../index.html" class="nav-link">首页</a></li>
                    <li class="nav-dropdown">
                        <a href="#" class="nav-link">行业方案 ▼</a>
                        <ul class="dropdown-menu">
                            <li><a href="../huawei-f5g.html">华为 F5G</a></li>
                            <li><a href="../../h3c-optical.html">华三 园区光网络</a></li>
                            <li><a href="../../zte-pon.html">中兴 PON方案</a></li>
                            <li><a href="../../fiberhome-pon.html">烽火 PON方案</a></li>
                        </ul>
                    </li>
                    <li><a href="../../../pages/news/index.html" class="nav-link">最新资讯</a></li>
                    <li><a href="../../../pages/docs/center.html" class="nav-link">技术文档</a></li>
                    <li><a href="../../../insights.html" class="nav-link">技术洞察</a></li>
                </ul>
            </div>
        </nav>'''
    
    pagination = ''
    if prev_link or next_link:
        pagination = '<div class="pagination">'
        if prev_link:
            pagination += f'<a href="{prev_link[0]}" class="page-link">← {prev_link[1]}</a>'
        pagination += f'<a href="index.html" class="page-link">目录</a>'
        if next_link:
            pagination += f'<a href="{next_link[0]}" class="page-link">{next_link[1]} →</a>'
        pagination += '</div>'
    
    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{title} | PON Portal</title>
    <link rel="stylesheet" href="../../../css/main.css">
    <link rel="stylesheet" href="../../../css/components.css">
    <link rel="stylesheet" href="../../../css/responsive.css">
    <style>
        .data-table {{
            width: 100%;
            border-collapse: collapse;
            background: white;
            border-radius: 8px;
            overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            font-size: 0.9rem;
        }}
        .data-table th {{
            background: linear-gradient(135deg, #1e3a8a, #3b82f6);
            color: white;
            padding: 12px 10px;
            text-align: left;
            font-weight: 600;
            white-space: nowrap;
        }}
        .data-table td {{
            padding: 10px;
            border-bottom: 1px solid #e2e8f0;
            max-width: 300px;
            overflow: hidden;
            text-overflow: ellipsis;
            white-space: nowrap;
        }}
        .data-table tr:hover {{
            background: #f8fafc;
        }}
        .data-table tr:nth-child(even) {{
            background: #f8fafc;
        }}
        .data-table tr:nth-child(even):hover {{
            background: #f1f5f9;
        }}
        .table-wrapper {{
            overflow-x: auto;
            margin: 20px 0;
        }}
        .sheet-header {{
            background: linear-gradient(135deg, #1e3a8a, #3b82f6);
            color: white;
            padding: 40px;
            border-radius: 12px;
            margin-bottom: 30px;
            text-align: center;
        }}
        .sheet-header h1 {{
            margin: 0 0 10px 0;
            font-size: 1.8rem;
        }}
        .sheet-header p {{
            margin: 0;
            opacity: 0.9;
        }}
        .breadcrumb {{
            padding: 15px 0;
            font-size: 0.9rem;
            color: #6b7280;
        }}
        .breadcrumb a {{
            color: #3b82f6;
            text-decoration: none;
        }}
        .breadcrumb a:hover {{
            text-decoration: underline;
        }}
        .download-section {{
            background: #f0fdf4;
            border: 1px solid #86efac;
            border-radius: 8px;
            padding: 20px;
            margin: 20px 0;
            text-align: center;
        }}
        .download-btn {{
            display: inline-block;
            background: #16a34a;
            color: white;
            padding: 12px 30px;
            border-radius: 8px;
            text-decoration: none;
            font-weight: 600;
            transition: background 0.2s;
        }}
        .download-btn:hover {{
            background: #15803d;
        }}
        .pagination {{
            display: flex;
            justify-content: space-between;
            align-items: center;
            margin: 30px 0;
            padding: 20px;
            background: #f8fafc;
            border-radius: 8px;
        }}
        .page-link {{
            padding: 10px 20px;
            background: #3b82f6;
            color: white;
            text-decoration: none;
            border-radius: 6px;
            font-weight: 500;
        }}
        .page-link:hover {{
            background: #1e40af;
        }}
        .section-info {{
            background: white;
            border-radius: 8px;
            padding: 20px;
            margin: 20px 0;
            border-left: 4px solid #3b82f6;
        }}
    </style>
</head>
<body>
    {nav}
    
    <main class="container" style="padding-top: 80px; padding-bottom: 60px;">
        <div class="breadcrumb">
            <a href="../../../index.html">首页</a> / 
            <a href="../huawei-f5g.html">华为 F5G</a> / 
            {index_link} / 
            {escape_html(title)}
        </div>
        
        <div class="sheet-header">
            <h1>{sheet_info['icon']} {escape_html(title)}</h1>
            <p>数据来源：POL竞争分析.xlsx</p>
        </div>
        
        <div class="download-section">
            <p><strong>📥 原始文件下载</strong></p>
            <a href="POL竞争分析.xlsx" class="download-btn" download>下载 POL竞争分析.xlsx</a>
        </div>
        
        <div class="section-info">
            <strong>📊 数据概览：</strong>该页面数据来自 Excel 文件中的 "{escape_html(title)}" 工作表
        </div>
        
        {content}
        
        {pagination}
    </main>
    
    <footer style="text-align: center; padding: 30px; color: #6b7280; font-size: 0.9rem;">
        <p>PON Network Portal | 行业分析 - 华为</p>
    </footer>
</body>
</html>'''

def generate_index_page(sheets_list):
    """Generate index page listing all sheets"""
    nav = '''
    <nav class="navbar">
        <div class="nav-container">
            <a href="../../../index.html" class="nav-logo">
                <span class="logo-icon">🌐</span>
                <span class="logo-text">PON Portal</span>
            </a>
            <ul class="nav-menu" id="navMenu">
                <li><a href="../../../index.html" class="nav-link">首页</a></li>
                <li class="nav-dropdown">
                    <a href="#" class="nav-link">行业方案 ▼</a>
                    <ul class="dropdown-menu">
                        <li><a href="../huawei-f5g.html">华为 F5G</a></li>
                        <li><a href="../../h3c-optical.html">华三 园区光网络</a></li>
                        <li><a href="../../zte-pon.html">中兴 PON方案</a></li>
                        <li><a href="../../fiberhome-pon.html">烽火 PON方案</a></li>
                    </ul>
                </li>
                <li><a href="../../../pages/news/index.html" class="nav-link">最新资讯</a></li>
                <li><a href="../../../pages/docs/center.html" class="nav-link">技术文档</a></li>
                <li><a href="../../../insights.html" class="nav-link">技术洞察</a></li>
            </ul>
        </div>
    </nav>'''
    
    cards_html = []
    for i, (sheet_name, config) in enumerate(SHEET_CONFIG.items()):
        slug, display_name, icon = config
        prev_sheet = None
        next_sheet = None
        if i > 0:
            prev_name = list(SHEET_CONFIG.keys())[i-1]
            prev_config = SHEET_CONFIG[prev_name]
            prev_sheet = (prev_config[0] + '.html', prev_config[1])
        if i < len(SHEET_CONFIG) - 1:
            next_name = list(SHEET_CONFIG.keys())[i+1]
            next_config = SHEET_CONFIG[next_name]
            next_sheet = (next_config[0] + '.html', next_config[1])
        
        cards_html.append(f'''
                <div class="sheet-card">
                    <div class="card-icon">{icon}</div>
                    <h3>{escape_html(display_name)}</h3>
                    <p class="sheet-meta">Sheet {i+1} of {len(SHEET_CONFIG)}</p>
                    <a href="{slug}.html" class="card-link">查看详情 →</a>
                </div>''')
    
    cards_section = '\n'.join(cards_html)
    
    return f'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>POL竞争分析 - 华为 | PON Portal</title>
    <link rel="stylesheet" href="../../../css/main.css">
    <link rel="stylesheet" href="../../../css/components.css">
    <link rel="stylesheet" href="../../../css/responsive.css">
    <style>
        .page-header {{
            background: linear-gradient(135deg, #1e3a8a, #3b82f6);
            color: white;
            padding: 60px 40px;
            border-radius: 12px;
            text-align: center;
            margin-bottom: 40px;
        }}
        .page-header h1 {{
            margin: 0 0 15px 0;
            font-size: 2.2rem;
        }}
        .page-header p {{
            margin: 0;
            opacity: 0.9;
            font-size: 1.1rem;
        }}
        .breadcrumb {{
            padding: 15px 0;
            font-size: 0.9rem;
            color: #6b7280;
        }}
        .breadcrumb a {{
            color: #3b82f6;
            text-decoration: none;
        }}
        .breadcrumb a:hover {{
            text-decoration: underline;
        }}
        .sheet-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
            gap: 20px;
            margin: 30px 0;
        }}
        .sheet-card {{
            background: white;
            border-radius: 12px;
            padding: 25px;
            box-shadow: 0 2px 10px rgba(0,0,0,0.08);
            transition: transform 0.2s, box-shadow 0.2s;
            border-left: 4px solid #3b82f6;
        }}
        .sheet-card:hover {{
            transform: translateY(-3px);
            box-shadow: 0 4px 20px rgba(0,0,0,0.12);
        }}
        .sheet-card .card-icon {{
            font-size: 2.5rem;
            margin-bottom: 15px;
        }}
        .sheet-card h3 {{
            color: #1e3a8a;
            margin: 0 0 10px 0;
            font-size: 1.1rem;
        }}
        .sheet-meta {{
            color: #6b7280;
            font-size: 0.85rem;
            margin: 5px 0;
        }}
        .card-link {{
            display: inline-block;
            margin-top: 15px;
            padding: 8px 16px;
            background: #eff6ff;
            color: #1e40af;
            text-decoration: none;
            border-radius: 6px;
            font-size: 0.9rem;
            font-weight: 500;
            transition: background 0.2s;
        }}
        .card-link:hover {{
            background: #dbeafe;
        }}
        .download-section {{
            background: #f0fdf4;
            border: 1px solid #86efac;
            border-radius: 12px;
            padding: 30px;
            margin: 30px 0;
            text-align: center;
        }}
        .download-btn {{
            display: inline-block;
            background: #16a34a;
            color: white;
            padding: 14px 35px;
            border-radius: 8px;
            text-decoration: none;
            font-weight: 600;
            font-size: 1rem;
            transition: background 0.2s;
        }}
        .download-btn:hover {{
            background: #15803d;
        }}
        .info-box {{
            background: white;
            border-radius: 12px;
            padding: 25px;
            margin: 20px 0;
            border-left: 4px solid #3b82f6;
            box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        }}
    </style>
</head>
<body>
    {nav}
    
    <main class="container" style="padding-top: 80px; padding-bottom: 60px;">
        <div class="breadcrumb">
            <a href="../../../index.html">首页</a> / 
            <a href="../huawei-f5g.html">华为 F5G</a> / 
            行业分析
        </div>
        
        <div class="page-header">
            <h1>🏢 POL竞争分析 - 华为</h1>
            <p>华为PON网络竞争情报，包含市场竞争、产品规格、方案对比等15个分析维度</p>
        </div>
        
        <div class="download-section">
            <p style="margin: 0 0 15px 0; font-size: 1.1rem;"><strong>📥 下载原始文件</strong></p>
            <a href="POL竞争分析.xlsx" class="download-btn" download>下载 POL竞争分析.xlsx</a>
        </div>
        
        <div class="info-box">
            <h3 style="margin-top: 0;">📊 内容说明</h3>
            <p>本目录包含华为POL竞争分析的完整数据，涵盖以下内容：</p>
            <ul>
                <li><strong>市场竞争分析</strong>：华为、中兴、烽火、诺基亚、爱立信五大厂商对比</li>
                <li><strong>产品规格</strong>：OLT、ONU、MA5800、EA5800等设备详细规格</li>
                <li><strong>方案规格</strong>：eSight网管、F5G配置、保护组等技术规格</li>
                <li><strong>无线规格</strong>：POL无线能力对比、Wi-Fi特性分析</li>
            </ul>
        </div>
        
        <h2 style="color: #1e3a8a; margin: 30px 0 20px 0;">📑 选择分析维度</h2>
        <div class="sheet-grid">
            {cards_section}
        </div>
    </main>
    
    <footer style="text-align: center; padding: 30px; color: #6b7280; font-size: 0.9rem;">
        <p>PON Network Portal | 行业分析 - 华为</p>
    </footer>
</body>
</html>'''

def main():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    sheets_list = list(SHEET_CONFIG.items())
    
    # Generate index page
    index_html = generate_index_page(sheets_list)
    with open(os.path.join(BASE_DIR, 'index.html'), 'w', encoding='utf-8') as f:
        f.write(index_html)
    print(f'✓ Created index.html')
    
    # Generate individual sheet pages
    for i, (sheet_name, config) in enumerate(sheets_list):
        if sheet_name not in wb.sheetnames:
            print(f'⚠ Sheet "{sheet_name}" not found in workbook, skipping')
            continue
        
        slug, display_name, icon = config
        
        # Get prev/next links
        prev_info = None
        next_info = None
        if i > 0:
            prev_name = sheets_list[i-1][0]
            prev_config = sheets_list[i-1][1]
            prev_info = (prev_config[0] + '.html', prev_config[1])
        if i < len(sheets_list) - 1:
            next_name = sheets_list[i+1][0]
            next_config = sheets_list[i+1][1]
            next_info = (next_config[0] + '.html', next_config[1])
        
        sheet_info = {
            'icon': icon,
            'prev': prev_info,
            'next': next_info,
        }
        
        ws = wb[sheet_name]
        content = generate_table(ws)
        
        page_html = generate_page_template(display_name, content, '', sheet_info)
        
        output_path = os.path.join(BASE_DIR, f'{slug}.html')
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(page_html)
        print(f'✓ Created {slug}.html ({ws.max_row} rows)')
    
    print(f'\n✅ All pages generated in {BASE_DIR}')

if __name__ == '__main__':
    main()
